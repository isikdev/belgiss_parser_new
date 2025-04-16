import os
import sys
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import re
from colorama import init, Fore, Back, Style
import logging
import traceback
import uuid
import importlib.util
import importlib.machinery
import random
import json
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import requests
import urllib3
import concurrent.futures
import threading
import time
from tqdm import tqdm
import queue
import math

# Инициализация colorama для поддержки цветов в Windows
init(autoreset=True)

# Отключаем предупреждения для незащищенных запросов
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Настройка логирования
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir, exist_ok=True)

# Создаем уникальный ID сессии
SESSION_ID = str(uuid.uuid4())[:8]

# Настраиваем логгер
logging.basicConfig(
    filename=os.path.join(log_dir, f"details_excel_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{SESSION_ID}.log"),
    level=logging.DEBUG,
    format='%(asctime)s - [%(levelname)s] - %(message)s'
)

# Добавляем вывод логов в консоль
console = logging.StreamHandler()
console.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - [%(levelname)s] - %(message)s')
console.setFormatter(formatter)
logging.getLogger('').addHandler(console)

# Функции для логирования
def log_info(message):
    logging.info(message)

def log_error(message):
    logging.error(message)

def log_debug(message):
    logging.debug(message)

def log_warning(message):
    logging.warning(message)

# Создаем директории для сохранения данных
output_dir = "declarations_details"
os.makedirs(output_dir, exist_ok=True)

# Базовый URL API для деталей деклараций
base_url = "https://api.belgiss.by/tsouz/tsouz-certifs"

# Список заголовков для имитации браузера
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36 Edg/92.0.902.84"
]

# Глобальные переменные для прокси
proxy_list = []
proxy_lock = threading.Lock()
proxy_stats = {}
proxy_limiters = {}
current_proxy_index = 0

def clear_screen():
    """Очистка экрана терминала"""
    try:
        os.system('cls' if os.name == 'nt' else 'clear')
        log_debug("Экран очищен")
    except Exception as e:
        log_error(f"Ошибка при очистке экрана: {e}")

def print_header():
    """Выводит заголовок приложения"""
    log_debug("Вывод заголовка")
    clear_screen()
    print(Fore.CYAN + "=" * 80 + Style.RESET_ALL)
    print(Fore.CYAN + " " * 20 + Fore.YELLOW + Style.BRIGHT + "ЗАГРУЗКА ДЕТАЛЕЙ ДЕКЛАРАЦИЙ И ГЕНЕРАЦИЯ EXCEL" + Fore.CYAN + " " * 20 + Style.RESET_ALL)
    print(Fore.CYAN + "=" * 80 + Style.RESET_ALL)
    print("\n")

def print_footer():
    """Выводит подвал приложения"""
    log_debug("Вывод подвала")
    print(Fore.CYAN + "=" * 80 + Style.RESET_ALL)
    print(Fore.CYAN + "=" * 80 + Style.RESET_ALL)

def load_proxies(proxy_file=None):
    """Загружает список прокси-серверов из файла"""
    global proxy_list, proxy_stats, proxy_limiters
    
    if proxy_file and os.path.exists(proxy_file):
        encodings = ['utf-8', 'utf-16', 'utf-16-le', 'utf-16-be', 'cp1251', 'latin-1']
        
        for encoding in encodings:
            try:
                with open(proxy_file, 'r', encoding=encoding) as f:
                    loaded_proxies = []
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            loaded_proxies.append(line)
                    
                    proxy_list = loaded_proxies
                    
                    for proxy in proxy_list:
                        proxy_stats[proxy] = {
                            'success': 0,
                            'errors': 0,
                            'rate_limits': 0,
                            'last_error': None,
                            'is_active': True
                        }
                        proxy_limiters[proxy] = AdaptiveRateLimiter()
                    
                    log_info(f"Загружено {len(proxy_list)} прокси из файла {proxy_file}")
                    return True
            except UnicodeDecodeError:
                continue
            except Exception as e:
                log_error(f"Ошибка при загрузке прокси: {e}")
                return False
    
    return False

def get_random_headers():
    """Возвращает случайные заголовки для запроса"""
    return {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        'Referer': 'https://belgiss.by/',
        'Origin': 'https://belgiss.by'
    }

class AdaptiveRateLimiter:
    """Адаптивный ограничитель скорости запросов"""
    def __init__(self, initial_rate=1.0, max_rate=3.0, backoff_factor=0.5, recovery_factor=1.05):
        self.rate = initial_rate
        self.max_rate = max_rate
        self.backoff_factor = backoff_factor
        self.recovery_factor = recovery_factor
        self.last_calls = []
        self.lock = threading.Lock()

    def wait_for_permission(self):
        """Ожидает разрешения на выполнение запроса"""
        with self.lock:
            now = time.time()
            self._cleanup_old_calls()
            
            if len(self.last_calls) > 0:
                time_since_last = now - self.last_calls[-1]
                if time_since_last < 1.0 / self.rate:
                    time.sleep(1.0 / self.rate - time_since_last)
            
            self.last_calls.append(now)

    def report_success(self):
        """Увеличивает скорость при успешном запросе"""
        with self.lock:
            self.rate = min(self.rate * self.recovery_factor, self.max_rate)

    def report_error(self, is_rate_limit_error=True):
        """Уменьшает скорость при ошибке"""
        with self.lock:
            if is_rate_limit_error:
                self.rate = max(self.rate * self.backoff_factor, 0.1)

    def _cleanup_old_calls(self):
        """Удаляет старые записи о вызовах"""
        now = time.time()
        self.last_calls = [t for t in self.last_calls if now - t < 60]

def make_request_with_retry(url, proxy=None, max_retries=5, initial_delay=2.0, proxy_timeout=300):
    """Выполняет запрос с повторными попытками"""
    headers = get_random_headers()
    
    for attempt in range(max_retries):
        try:
            if proxy:
                proxies = {
                    'http': f'http://{proxy}',
                    'https': f'http://{proxy}'
                }
            else:
                proxies = None
            
            response = requests.get(
                url,
                headers=headers,
                proxies=proxies,
                verify=False,
                timeout=proxy_timeout
            )
            
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 429:  # Rate limit
                time.sleep(initial_delay * (2 ** attempt))
                continue
            else:
                log_error(f"Ошибка HTTP {response.status_code} для URL {url}")
                time.sleep(initial_delay * (2 ** attempt))
                
        except Exception as e:
            log_error(f"Ошибка при запросе {url}: {e}")
            time.sleep(initial_delay * (2 ** attempt))
    
    return None

def download_declaration_details(declaration_id, batch_folder, limiter):
    """Скачивает детали декларации"""
    url = f"{base_url}/{declaration_id}"
    
    try:
        limiter.wait_for_permission()
        
        # Проверяем, существует ли уже файл
        filename = os.path.join(batch_folder, f"{declaration_id}.json")
        if os.path.exists(filename):
            return True
            
        max_retries = 5
        retry_delay = 2
        
        for attempt in range(max_retries):
            try:
                data = make_request_with_retry(url)
                if data:
                    with open(filename, 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                    return True
                elif attempt < max_retries - 1:
                    time.sleep(retry_delay * (attempt + 1))
            except Exception:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (attempt + 1))
                continue
        
        return False
        
    except Exception:
        return False

def load_declarations_from_json(file_path):
    """Загружает список деклараций из JSON файла"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            items = data.get("items", [])
            return [item.get("certdecltr_id") for item in items if "certdecltr_id" in item]
    except Exception as e:
        log_error(f"Ошибка при загрузке файла {file_path}: {e}")
        return []

def scan_directory_for_json(directory):
    """Сканирует директорию на наличие JSON файлов с декларациями"""
    log_info(f"Начало сканирования директории: {directory}")
    json_files = []
    try:
        for root, _, files in os.walk(directory):
            for file in files:
                if file.endswith('.json') and not file.endswith('download_report.json'):
                    file_path = os.path.join(root, file)
                    json_files.append(file_path)
                    log_debug(f"Найден JSON файл: {file_path}")
        log_info(f"Сканирование завершено. Найдено {len(json_files)} файлов")
    except Exception as e:
        log_error(f"Ошибка при сканировании директории {directory}: {e}")
        log_error(traceback.format_exc())
    return json_files

def format_time(seconds):
    """Форматирует время в читаемый вид"""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def download_all_declaration_details(declaration_files, workers=50, resume=False, batch_size=500, 
                                  initial_delay=2.0, max_retries=5, proxy_timeout=300):
    """Скачивает все детали деклараций"""
    # Отключаем вывод логов в консоль
    logging.getLogger().handlers = []
    logging.getLogger().addHandler(logging.NullHandler())
    
    # Собираем все ID деклараций из файлов
    all_declaration_ids = []
    for file_path in declaration_files:
        try:
            ids = load_declarations_from_json(file_path)
            all_declaration_ids.extend(ids)
        except Exception as e:
            continue
    
    if not all_declaration_ids:
        print(Fore.RED + "\nНе найдено ID деклараций в файлах" + Style.RESET_ALL)
        return 0, 0
    
    # Удаляем дубликаты и перемешиваем
    all_declaration_ids = list(set(all_declaration_ids))
    random.shuffle(all_declaration_ids)
    
    # Проверяем уже загруженные декларации
    if resume:
        completed_ids = set()
        for root, _, files in os.walk(output_dir):
            for file in files:
                if file.endswith('.json') and not file == "download_report.json":
                    try:
                        declaration_id = int(file.split('.')[0])
                        completed_ids.add(declaration_id)
                    except ValueError:
                        continue
        
        # Удаляем уже загруженные ID из списка
        all_declaration_ids = [id for id in all_declaration_ids if id not in completed_ids]
        if len(completed_ids) > 0:
            print(Fore.GREEN + f"\nПропускаем {len(completed_ids)} уже загруженных деклараций" + Style.RESET_ALL)
    
    total_declarations = len(all_declaration_ids)
    if total_declarations == 0:
        print(Fore.GREEN + "\nВсе декларации уже загружены!" + Style.RESET_ALL)
        return 0, 0
    
    # Переменные для отображения статистики
    completed = 0
    success = 0
    errors = 0
    start_time = time.time()
    
    try:
        # Создаем батчи из ID
        batches = [all_declaration_ids[i:i + batch_size] for i in range(0, total_declarations, batch_size)]
        
        for batch_index, batch_ids in enumerate(batches, 1):
            batch_folder = os.path.join(output_dir, f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{batch_index}")
            os.makedirs(batch_folder, exist_ok=True)
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
                futures = {executor.submit(download_declaration_details, declaration_id, batch_folder, AdaptiveRateLimiter()): declaration_id 
                          for declaration_id in batch_ids}
                
                for future in concurrent.futures.as_completed(futures):
                    try:
                        if future.result():
                            success += 1
                        else:
                            errors += 1
                        
                        completed += 1
                        
                        # Обновляем статус на одной строке
                        elapsed_time = time.time() - start_time
                        speed = completed / elapsed_time if elapsed_time > 0 else 0
                        remaining = ((total_declarations - completed) / speed) if speed > 0 else 0
                        
                        # Очищаем предыдущую строку
                        print("\r" + " " * 100, end="\r")
                        
                        # Создаем строку прогресса
                        progress = (completed / total_declarations) * 100
                        progress_bar = "=" * int(progress // 2) + ">" + " " * (50 - int(progress // 2))
                        status = f"Прогресс: [{progress_bar}] {progress:.1f}% ({completed}/{total_declarations}) | "
                        status += f"Успешно: {success} | Ошибки: {errors} | "
                        status += f"Скорость: {speed:.1f} декл/сек | "
                        status += f"Осталось: {format_time(remaining)}"
                        
                        print(f"\r{Fore.CYAN}{status}{Style.RESET_ALL}", end="")
                        
                    except Exception:
                        errors += 1
                        completed += 1
            
            # Делаем паузу между батчами
            if batch_index < len(batches):
                time.sleep(2)
    
    except KeyboardInterrupt:
        print(Fore.YELLOW + "\n\nЗагрузка прервана пользователем. Прогресс сохранен." + Style.RESET_ALL)
    except Exception as e:
        print(Fore.RED + f"\n\nКритическая ошибка: {e}" + Style.RESET_ALL)
    
    print("\n")  # Перевод строки после завершения
    return success, errors

def get_value_safely(data, *keys, default=""):
    """Безопасно извлекает значение из вложенного словаря по цепочке ключей"""
    current = data
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        elif isinstance(current, list) and isinstance(key, int) and len(current) > key:
            current = current[key]
        else:
            return default
    
    # Преобразуем None в пустую строку
    if current is None:
        return default
    
    # Если получили список, возвращаем первый элемент или склеиваем все элементы
    if isinstance(current, list):
        if len(current) == 0:
            return default
        # Если список объектов, пытаемся извлечь полезную информацию из первого элемента
        elif isinstance(current[0], dict):
            if len(current) == 1:
                # Если в словаре есть понятные значения, возвращаем их
                for key in ['Name', 'Description', 'Text', 'Value', 'Id']:
                    if key in current[0]:
                        return str(current[0][key])
                # Иначе возвращаем весь словарь как строку
                return str(current[0])
            else:
                # При множестве элементов возвращаем список строковых представлений
                return ", ".join(str(item) for item in current)
        elif len(current) == 1:
            return str(current[0])
        else:
            return ", ".join(str(item) for item in current)
    
    # Преобразуем результат в строку
    return str(current)

def parse_date(date_str):
    """Пытается разобрать дату из строки в различных форматах"""
    if not date_str:
        return ""
    
    # Если date_str не строка, преобразуем в строку
    if not isinstance(date_str, str):
        date_str = str(date_str)
    
    formats = [
        '%d.%m.%Y',  # 01.01.2020
        '%Y-%m-%d',  # 2020-01-01
        '%Y-%m-%dT%H:%M:%S',  # 2020-01-01T00:00:00
        '%Y/%m/%d',  # 2020/01/01
        '%d-%m-%Y',  # 01-01-2020
    ]
    
    for fmt in formats:
        try:
            date_obj = datetime.strptime(date_str, fmt)
            # Возвращаем в формате ДД.ММ.ГГГГ
            return date_obj.strftime('%d.%m.%Y')
        except ValueError:
            continue
    
    # Если не удалось разобрать дату стандартными форматами, пробуем извлечь ее из строки
    # Ищем шаблон ДД.ММ.ГГГГ
    match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', date_str)
    if match:
        try:
            day, month, year = map(int, match.groups())
            return f"{day:02d}.{month:02d}.{year:04d}"
        except ValueError:
            pass
    
    # Если все методы не сработали, возвращаем исходную строку
    return date_str

def process_declaration_details(json_data):
    """Обрабатывает детали декларации из JSON и возвращает словарь с данными"""
    try:
        # Проверка валидности данных
        if not isinstance(json_data, dict):
            log_error("Неверный формат данных JSON")
            return None
        
        # Находим основные данные документа
        doc_data = json_data.get("certdecltr_ConformityDocDetails", {})
        
        # Регистрационный номер
        doc_number = get_value_safely(doc_data, "DocId")
        
        # Ссылка на документ
        doc_id = json_data.get("certdecltr_id") or json_data.get("documents_id")
        doc_link = f"https://tsouz.belgiss.by/doc.php?id={doc_id}" if doc_id else ""
        
        # Статус действия сертификата
        doc_status = ""
        if "DocStatusDetails" in doc_data:
            status_code = get_value_safely(doc_data, "DocStatusDetails", "DocStatusCode")
            if status_code == "01":
                doc_status = "действует"
            elif status_code == "03":
                doc_status = "прекращен"
            elif status_code == "02":
                doc_status = "приостановлен"
        
        # Вид документа об оценке соответствия
        doc_type = ""
        conformity_kind_code = get_value_safely(doc_data, "ConformityDocKindCode")
        if conformity_kind_code == "10":
            doc_type = "Декларация о соответствии"
        elif conformity_kind_code == "1":
            doc_type = "Сертификат соответствия"
        
        # Номер технического регламента
        tech_regs = get_value_safely(doc_data, "TechnicalRegulationId", default=[])
        tech_reg = ", ".join(tech_regs) if isinstance(tech_regs, list) else tech_regs
        
        # Полное наименование органа по сертификации
        cert_org = get_value_safely(doc_data, "ConformityAuthorityV2Details", "BusinessEntityName")
        
        # Данные заявителя
        applicant = doc_data.get("ApplicantDetails", {})
        applicant_country = get_value_safely(applicant, "UnifiedCountryCode")
        applicant_name = get_value_safely(applicant, "BusinessEntityName")
        applicant_short_name = get_value_safely(applicant, "BusinessEntityBriefName")
        applicant_id = get_value_safely(applicant, "BusinessEntityId")
        
        # Адрес заявителя
        applicant_address = ""
        if "SubjectAddressDetails" in applicant and isinstance(applicant["SubjectAddressDetails"], list) and applicant["SubjectAddressDetails"]:
            addr = applicant["SubjectAddressDetails"][0]
            parts = [
                get_value_safely(addr, "RegionName"),
                get_value_safely(addr, "CityName"),
                get_value_safely(addr, "StreetName"),
                get_value_safely(addr, "BuildingNumberId")
            ]
            applicant_address = ", ".join(filter(None, parts))
        
        # Контактные данные заявителя
        applicant_contact = ""
        if "CommunicationDetails" in applicant and isinstance(applicant["CommunicationDetails"], list):
            contacts = []
            for comm in applicant["CommunicationDetails"]:
                if isinstance(comm, dict) and "CommunicationChannelId" in comm:
                    channel_id = comm["CommunicationChannelId"]
                    if isinstance(channel_id, list):
                        contacts.extend(channel_id)
                    else:
                        contacts.append(channel_id)
            applicant_contact = ", ".join(filter(None, contacts))
        
        # Данные изготовителя
        manufacturer = {}
        if "ManufacturerDetails" in doc_data and isinstance(doc_data["ManufacturerDetails"], list) and doc_data["ManufacturerDetails"]:
            manufacturer = doc_data["ManufacturerDetails"][0]
        
        manufacturer_country = get_value_safely(manufacturer, "UnifiedCountryCode")
        manufacturer_name = get_value_safely(manufacturer, "BusinessEntityBriefName") or get_value_safely(manufacturer, "BusinessEntityName")
        
        # Адрес изготовителя
        manufacturer_address = ""
        if "AddressV4Details" in manufacturer and isinstance(manufacturer["AddressV4Details"], list) and manufacturer["AddressV4Details"]:
            addr = manufacturer["AddressV4Details"][0]
            address_text = get_value_safely(addr, "AddressText")
            if address_text:
                manufacturer_address = address_text
            else:
                parts = [
                    get_value_safely(addr, "RegionName"),
                    get_value_safely(addr, "CityName"),
                    get_value_safely(addr, "StreetName"),
                    get_value_safely(addr, "BuildingNumberId")
                ]
                manufacturer_address = ", ".join(filter(None, parts))
        
        # Контактные данные изготовителя
        manufacturer_contact = ""
        if "CommunicationDetails" in manufacturer and isinstance(manufacturer["CommunicationDetails"], list):
            contacts = []
            for comm in manufacturer["CommunicationDetails"]:
                if isinstance(comm, dict) and "CommunicationChannelId" in comm:
                    channel_id = comm["CommunicationChannelId"]
                    if isinstance(channel_id, list):
                        contacts.extend(channel_id)
                    else:
                        contacts.append(channel_id)
            manufacturer_contact = ", ".join(filter(None, contacts))
        
        # Наименование объекта оценки соответствия
        product_name = get_value_safely(doc_data, "TechnicalRegulationObjectDetails", "ProductDetails", 0, "ProductName")
        
        # Код товара по ТН ВЭД ЕАЭС
        commodity_codes = get_value_safely(doc_data, "TechnicalRegulationObjectDetails", "ProductDetails", 0, "CommodityCode", default=[])
        commodity_code = ", ".join(commodity_codes) if isinstance(commodity_codes, list) else commodity_codes
        
        # Дополнительное наименование объекта
        product_text = get_value_safely(doc_data, "TechnicalRegulationObjectDetails", "ProductDetails", 0, "ProductText")
        
        # Даты документа
        doc_date = get_value_safely(doc_data, "DocStartDate")
        start_date = get_value_safely(doc_data, "DocStatusDetails", "StartDate") or doc_date
        end_date = get_value_safely(doc_data, "DocStatusDetails", "EndDate") or get_value_safely(doc_data, "DocValidityDate")
        
        # Форматируем даты
        doc_date = parse_date(doc_date)
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        
        return {
            "Регистрационный номер": doc_number,
            "Ссылка на документ": doc_link,
            "Статус действия сертификата (декларации)": doc_status,
            "Вид документа об оценке соответствия": doc_type,
            "Номер технического регламента": tech_reg,
            "Полное наименование органа по сертификации": cert_org,
            "Заявитель Страна": applicant_country,
            "Заявитель Краткое наименование": applicant_short_name or applicant_name,
            "Заявитель Идентификатор хозяйствующего субъекта": applicant_id,
            "Заявитель Адрес": applicant_address,
            "Заявитель Контактный реквизит": applicant_contact,
            "Изготовитель Страна": manufacturer_country,
            "Изготовитель Краткое наименование": manufacturer_name,
            "Изготовитель Адрес": manufacturer_address,
            "Изготовитель Контактный реквизит": manufacturer_contact,
            "Наименование объекта оценки соответствия": product_name,
            "Код товара по ТН ВЭД ЕАЭС": commodity_code,
            "Наименование объекта оценки соответствия (дополнительно)": product_text,
            "Дата документа": doc_date,
            "Дата начала действия": start_date,
            "Дата окончания действия": end_date
        }
    except Exception as e:
        log_error(f"Ошибка при обработке декларации: {str(e)}")
        return None

def format_excel_worksheet(worksheet, df):
    """Форматирует лист Excel согласно требованиям"""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Словарь с нужной шириной для каждой колонки
    column_widths = {
        'Регистрационный номер': 30,
        'Ссылка на документ': 35,
        'Статус действия сертификата (декларации)': 15,
        'Вид документа об оценке соответствия': 30,
        'Номер технического регламента': 25,
        'Полное наименование органа по сертификации': 40,
        'Заявитель Страна': 10,
        'Заявитель Краткое наименование': 40,
        'Заявитель Идентификатор хозяйствующего субъекта': 20,
        'Заявитель Адрес': 40,
        'Заявитель Контактный реквизит': 30,
        'Изготовитель Страна': 10,
        'Изготовитель Краткое наименование': 40,
        'Изготовитель Адрес': 40,
        'Изготовитель Контактный реквизит': 30,
        'Наименование объекта оценки соответствия': 40,
        'Код товара по ТН ВЭД ЕАЭС': 25,
        'Наименование объекта оценки соответствия (дополнительно)': 40,
        'Дата документа': 15,
        'Дата начала действия': 15,
        'Дата окончания действия': 15
    }
    
    # Устанавливаем ширину колонок
    for idx, col in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        width = column_widths.get(col, 20)  # По умолчанию 20, если колонка не указана в словаре
        worksheet.column_dimensions[column_letter].width = width
    
    # Заголовочный стиль
    header_font = Font(bold=True, size=11, name='Arial')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Границы
    medium_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Чередующиеся заливки строк
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Применяем форматирование к заголовкам
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = medium_border
    
    # Стиль для обычных ячеек
    cell_font = Font(name='Arial', size=10)
    
    # Применяем форматирование к данным (с чередующейся заливкой)
    for row in range(2, len(df) + 2):  # +2 потому что строка 1 - это заголовок, а df индексируется с 0
        # Определяем заполнение на основе четности строки
        row_fill = light_fill if row % 2 == 0 else None
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = cell_font
            
            # Разное выравнивание в зависимости от типа данных
            if df.columns[col-1] in ['Дата документа', 'Дата начала действия', 'Дата окончания действия', 
                                   'Заявитель Страна', 'Изготовитель Страна']:
                # Даты и короткие поля - по центру
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif df.columns[col-1] in ['Регистрационный номер', 'Номер технического регламента',
                                      'Код товара по ТН ВЭД ЕАЭС']:
                # Номера - выравнивание по левому краю без переноса
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                # Текстовые поля - по левому краю с переносом
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Применяем заливку, если нужно
            if row_fill:
                cell.fill = row_fill
            
            cell.border = thin_border
            
            # Добавляем гиперссылку для колонки "Ссылка на документ"
            if col == df.columns.get_loc("Ссылка на документ") + 1:
                if cell.value and isinstance(cell.value, str) and (cell.value.startswith('http://') or cell.value.startswith('https://')):
                    cell.hyperlink = cell.value
                    cell.font = Font(name='Arial', size=10, color='0000FF', underline='single')
    
    # Форматируем строки с подходящими статусами
    status_col_idx = None
    for idx, col_name in enumerate(df.columns, 1):
        if col_name == 'Статус действия сертификата (декларации)':
            status_col_idx = idx
            break
    
    if status_col_idx:
        for row in range(2, len(df) + 2):
            status_cell = worksheet.cell(row=row, column=status_col_idx)
            if status_cell.value:
                if status_cell.value.lower() == 'действует':
                    # Зеленый для действующих
                    status_cell.font = Font(name='Arial', size=10, color='006100')
                elif status_cell.value.lower() in ['прекращен', 'приостановлен']:
                    # Красный для прекращенных/приостановленных
                    status_cell.font = Font(name='Arial', size=10, color='9C0006')
    
    # Фиксируем первую строку
    worksheet.freeze_panes = worksheet.cell(row=2, column=1)

def process_declaration_files(files, batch_size=1000, output_dir="batch_results", debug_mode=False, resume=False):
    """Обрабатывает файлы деклараций и создает Excel файлы"""
    try:
        os.makedirs(output_dir, exist_ok=True)
        
        # Разбиваем файлы на батчи
        batches = [files[i:i + batch_size] for i in range(0, len(files), batch_size)]
        
        batch_files = []
        for batch_index, batch in enumerate(batches, 1):
            batch_file = os.path.join(output_dir, f"batch_{batch_index:04d}.xlsx")
            
            # Пропускаем существующие батчи в режиме возобновления
            if resume and os.path.exists(batch_file):
                batch_files.append(batch_file)
                print(f"  Пропуск существующего батча: {batch_file}")
                continue
            
            declarations_data = []
            print(f"  Обработка батча {batch_index}/{len(batches)} ({len(batch)} файлов)")
            
            for i, file_path in enumerate(batch):
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    declaration_data = process_declaration_details(data)
                    if declaration_data:
                        declarations_data.append(declaration_data)
                    
                    # Отображаем прогресс каждые 100 файлов
                    if (i + 1) % 100 == 0 or i + 1 == len(batch):
                        print(f"    Прогресс: {i + 1}/{len(batch)} ({(i + 1) / len(batch) * 100:.1f}%)")
                except Exception as e:
                    log_error(f"Ошибка при обработке файла {file_path}: {e}")
                    continue
            
            if declarations_data:
                # Создаем DataFrame
                df = pd.DataFrame(declarations_data)
                
                # Определяем порядок колонок
                columns_order = [
                    'Регистрационный номер', 
                    'Ссылка на документ', 
                    'Статус действия сертификата (декларации)', 
                    'Вид документа об оценке соответствия',
                    'Номер технического регламента',
                    'Полное наименование органа по сертификации',
                    'Заявитель Страна', 
                    'Заявитель Краткое наименование', 
                    'Заявитель Идентификатор хозяйствующего субъекта',
                    'Заявитель Адрес', 
                    'Заявитель Контактный реквизит',
                    'Изготовитель Страна', 
                    'Изготовитель Краткое наименование', 
                    'Изготовитель Адрес',
                    'Изготовитель Контактный реквизит',
                    'Наименование объекта оценки соответствия',
                    'Код товара по ТН ВЭД ЕАЭС',
                    'Наименование объекта оценки соответствия (дополнительно)',
                    'Дата документа',
                    'Дата начала действия',
                    'Дата окончания действия'
                ]
                
                # Переупорядочиваем колонки (только те, которые существуют в df)
                existing_columns = [col for col in columns_order if col in df.columns]
                df = df.reindex(columns=existing_columns)
                
                print(f"  Сохранение батча {batch_index} в файл {batch_file}...")
                
                # Сохраняем в Excel
                with pd.ExcelWriter(batch_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Declarations')
                    worksheet = writer.sheets['Declarations']
                    format_excel_worksheet(worksheet, df)
                
                batch_files.append(batch_file)
                print(f"  Батч {batch_index} сохранен ({len(declarations_data)} записей)")
            else:
                print(f"  ВНИМАНИЕ: Батч {batch_index} не содержит данных для сохранения")
        
        return batch_files
        
    except Exception as e:
        log_error(f"Критическая ошибка при обработке файлов: {e}")
        log_error(traceback.format_exc())
        raise

def merge_batch_results(batch_files, output_file="declarations_final.xlsx"):
    """Объединяет результаты всех батчей в один файл"""
    if not batch_files:
        print("Нет файлов для объединения")
        return False
    
    print(f"Объединение {len(batch_files)} файлов в итоговый отчет...")
    
    # Список для всех данных деклараций
    all_data = []
    
    # Загружаем данные из каждого батч-файла
    for i, batch_file in enumerate(batch_files):
        print(f"  Загрузка данных из файла {i+1}/{len(batch_files)}: {batch_file}")
        
        try:
            # Читаем Excel без автоматического преобразования типов
            df = pd.read_excel(batch_file, parse_dates=False)
            print(f"    Загружено {len(df)} записей")
            
            # Преобразуем DataFrame в список словарей и добавляем в общий список
            batch_data = df.to_dict('records')
            all_data.extend(batch_data)
            
        except Exception as e:
            print(f"    Ошибка при загрузке файла {batch_file}: {str(e)}")
    
    print(f"Всего загружено {len(all_data)} записей из {len(batch_files)} файлов")
    
    # Создаем итоговый отчет
    if all_data:
        print(f"Создание итогового файла {output_file}...")
        
        # Создаем DataFrame из всех данных
        df = pd.DataFrame(all_data)
        
        # Определяем порядок колонок (такой же, как в generate_declarations_excel.py)
        columns_order = [
            'Регистрационный номер', 
            'Ссылка на документ', 
            'Статус действия сертификата (декларации)', 
            'Вид документа об оценке соответствия',
            'Номер технического регламента',
            'Полное наименование органа по сертификации',
            'Заявитель Страна', 
            'Заявитель Краткое наименование', 
            'Заявитель Идентификатор хозяйствующего субъекта',
            'Заявитель Адрес', 
            'Заявитель Контактный реквизит',
            'Изготовитель Страна', 
            'Изготовитель Краткое наименование', 
            'Изготовитель Адрес',
            'Изготовитель Контактный реквизит',
            'Наименование объекта оценки соответствия',
            'Код товара по ТН ВЭД ЕАЭС',
            'Наименование объекта оценки соответствия (дополнительно)',
            'Дата документа',
            'Дата начала действия',
            'Дата окончания действия'
        ]
        
        # Переупорядочиваем колонки (только те, которые существуют в df)
        existing_columns = [col for col in columns_order if col in df.columns]
        df = df.reindex(columns=existing_columns)
        
        # Сохраняем в Excel с форматированием
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Declarations')
            worksheet = writer.sheets['Declarations']
            format_excel_worksheet(worksheet, df)
        
        print(f"Итоговый файл успешно создан: {output_file}")
        return True
    else:
        print("Ошибка: не удалось загрузить данные из промежуточных файлов")
        return False

def main():
    log_info(f"====== ЗАПУСК ПРОГРАММЫ (Сессия: {SESSION_ID}) ======")
    
    try:
        print_header()
        log_info("Заголовок программы выведен")
        
        # Шаг 1: Запрос о наличии прокси
        log_info("Шаг 1: Запрос о наличии прокси")
        proxy_file = None
        while True:
            print(Fore.GREEN + "Есть ли у вас файл с прокси?" + Style.RESET_ALL)
            print(Fore.YELLOW + "1. " + Fore.WHITE + "Да")
            print(Fore.YELLOW + "2. " + Fore.WHITE + "Нет")
            log_debug("Ожидание ввода пользователя (1/2)")
            
            try:
                proxy_choice = input(Fore.CYAN + "\nВаш выбор (1/2): " + Style.RESET_ALL).strip()
                log_debug(f"Получен ввод: '{proxy_choice}'")
                
                if proxy_choice == "1":
                    log_info("Пользователь выбрал использование прокси-файла")
                    print(Fore.WHITE + "\nПожалуйста, выберите файл с прокси в открывшемся окне..." + Style.RESET_ALL)
                    
                    try:
                        proxy_file = filedialog.askopenfilename(
                            title="Выберите файл с прокси",
                            filetypes=[("Текстовые файлы", "*.txt"), ("Все файлы", "*.*")]
                        )
                        log_debug(f"Выбор файла: '{proxy_file}'")
                        
                        if not proxy_file:
                            log_warning("Файл не был выбран")
                            print(Fore.RED + "\nФайл не выбран. Попробуйте снова.\n" + Style.RESET_ALL)
                            continue
                        
                        log_info(f"Выбран файл с прокси: {proxy_file}")
                        print(Fore.GREEN + f"\nВыбран файл: {proxy_file}" + Style.RESET_ALL)
                        break
                    except Exception as e:
                        log_error(f"Ошибка при выборе файла: {e}")
                        log_error(traceback.format_exc())
                        print(Fore.RED + f"\nОшибка при выборе файла: {e}\n" + Style.RESET_ALL)
                        continue
                        
                elif proxy_choice == "2":
                    log_info("Пользователь выбрал работу без прокси")
                    print(Fore.YELLOW + "\nПродолжаем без использования прокси." + Style.RESET_ALL)
                    break
                else:
                    log_warning(f"Некорректный ввод: '{proxy_choice}'")
                    print(Fore.RED + "\nНеверный выбор. Пожалуйста, введите 1 или 2.\n" + Style.RESET_ALL)
            except Exception as e:
                log_error(f"Ошибка при обработке ввода: {e}")
                log_error(traceback.format_exc())
                print(Fore.RED + f"\nОшибка при вводе: {e}" + Style.RESET_ALL)
        
        # Шаг 2: Загрузка прокси, если выбран файл
        if proxy_file:
            log_info("Начало загрузки прокси")
            if not load_proxies(proxy_file):
                log_error("Ошибка при загрузке прокси")
                print(Fore.RED + "\nОшибка при загрузке прокси. Продолжаем без них." + Style.RESET_ALL)
                proxy_file = None
            else:
                log_info("Прокси успешно загружены")
        
        # Шаг 3: Запрос папки с декларациями
        print_header()
        print(Fore.GREEN + "Выберите папку с декларациями для обработки..." + Style.RESET_ALL)
        
        try:
            declarations_dir = filedialog.askdirectory(
                title="Выберите папку с декларациями"
            )
            
            if not declarations_dir:
                log_error("Папка не выбрана")
                print(Fore.RED + "\nПапка не выбрана. Программа завершена." + Style.RESET_ALL)
                return
            
            log_info(f"Выбрана папка с декларациями: {declarations_dir}")
            print(Fore.GREEN + f"\nВыбрана папка: {declarations_dir}" + Style.RESET_ALL)
            
        except Exception as e:
            log_error(f"Ошибка при выборе папки: {e}")
            log_error(traceback.format_exc())
            print(Fore.RED + f"\nОшибка при выборе папки: {e}" + Style.RESET_ALL)
            return
        
        # Шаг 4: Загрузка и обработка деклараций
        print_header()
        print(Fore.GREEN + "Загрузка деклараций..." + Style.RESET_ALL)
        
        # Сканируем папку на наличие JSON файлов
        log_info("Начало сканирования папки с декларациями")
        declaration_files = scan_directory_for_json(declarations_dir)
        if not declaration_files:
            log_error("Не найдены файлы деклараций")
            print(Fore.RED + "\nНе найдены файлы деклараций в выбранной папке." + Style.RESET_ALL)
            return
        
        log_info(f"Найдено {len(declaration_files)} файлов деклараций")
        print(Fore.GREEN + f"\nНайдено {len(declaration_files)} файлов деклараций для обработки." + Style.RESET_ALL)
        
        # Запускаем загрузку деталей
        log_info("Начало загрузки деталей деклараций")
        try:
            success, errors = download_all_declaration_details(
                declaration_files,
                workers=50,
                resume=True,  # Всегда используем режим возобновления
                batch_size=500,
                initial_delay=2.0,
                max_retries=5,
                proxy_timeout=300
            )
            log_info(f"Загрузка завершена. Успешно: {success}, Ошибок: {errors}")
            print(Fore.GREEN + f"\nЗагрузка завершена. Успешно: {success}, Ошибок: {errors}" + Style.RESET_ALL)
        except Exception as e:
            log_error(f"Ошибка при загрузке деклараций: {e}")
            log_error(traceback.format_exc())
            print(Fore.RED + f"\nОшибка при загрузке деклараций: {e}" + Style.RESET_ALL)
            return
        
        # Шаг 5: Обработка загруженных деклараций
        print_header()
        print(Fore.GREEN + "Обработка загруженных деклараций..." + Style.RESET_ALL)
        
        log_info("Начало обработки загруженных деклараций")
        declaration_files = scan_directory_for_json(output_dir)
        if not declaration_files:
            log_error("Не найдены файлы деклараций для обработки")
            print(Fore.RED + "\nНе найдены файлы деклараций для обработки." + Style.RESET_ALL)
            return
        
        log_info(f"Найдено {len(declaration_files)} файлов для обработки")
        print(Fore.GREEN + f"\nНайдено {len(declaration_files)} файлов деклараций для обработки." + Style.RESET_ALL)
        
        # Обрабатываем файлы и создаем Excel
        log_info("Начало создания Excel файлов")
        try:
            batch_files = process_declaration_files(
                declaration_files,
                batch_size=1000,
                output_dir="batch_results",
                debug_mode=False,
                resume=True  # Всегда используем режим возобновления
            )
            
            if batch_files:
                log_info("Начало объединения результатов")
                if merge_batch_results(batch_files, "declarations_final.xlsx"):
                    log_info("Обработка успешно завершена")
                    print(Fore.GREEN + "\nОбработка успешно завершена!" + Style.RESET_ALL)
                else:
                    log_error("Ошибка при объединении результатов")
                    print(Fore.RED + "\nОшибка при объединении результатов." + Style.RESET_ALL)
            else:
                log_error("Не удалось создать файлы с результатами")
                print(Fore.RED + "\nНе удалось создать файлы с результатами." + Style.RESET_ALL)
        except Exception as e:
            log_error(f"Ошибка при обработке файлов: {e}")
            log_error(traceback.format_exc())
            print(Fore.RED + f"\nОшибка при обработке файлов: {e}" + Style.RESET_ALL)
            return
        
        print_footer()
        log_info("Программа успешно завершена")
        input(Fore.YELLOW + "\nНажмите Enter для завершения работы..." + Style.RESET_ALL)
        
    except Exception as e:
        log_error(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
        log_error(traceback.format_exc())
        print(Fore.RED + f"\nКРИТИЧЕСКАЯ ОШИБКА: {e}" + Style.RESET_ALL)
        input(Fore.RED + "\nНажмите Enter для завершения работы..." + Style.RESET_ALL)
        sys.exit(1)

if __name__ == "__main__":
    main() 