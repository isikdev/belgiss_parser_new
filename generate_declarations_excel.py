import os
import json
import glob
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re
import argparse


def extract_date_from_filename(filename):
    """Извлекает дату из имени файла или директории с батчем"""
    # Ищем дату в формате YYYYMMDD в имени директории батча
    match = re.search(r'batch_(\d{8})_', filename)
    if match:
        date_str = match.group(1)
        return datetime.strptime(date_str, '%Y%m%d')
    
    # Если не удалось найти дату в имени директории, возвращаем минимальную дату
    return datetime.min

def extract_date_from_json(json_data):
    """Извлекает дату декларации из данных JSON"""
    # Проверяем различные поля, которые могут содержать дату
    date_fields = [
        'DocStartDate',
        'DocIssueDate',
        'DocEndDate',
        'IncludeDate'
    ]
    
    for field in date_fields:
        if field in json_data and json_data[field]:
            try:
                # Пробуем разные форматы даты
                date_formats = ['%d.%m.%Y', '%Y-%m-%d', '%Y-%m-%dT%H:%M:%S']
                for fmt in date_formats:
                    try:
                        return datetime.strptime(json_data[field], fmt)
                    except ValueError:
                        continue
            except Exception:
                continue
    
    # Если не удалось извлечь дату из JSON, возвращаем минимальную дату
    return datetime.min


def find_declaration_files(base_dir="declarations_details"):
    """Находит все JSON файлы деклараций в указанной директории"""
    print("Поиск файлов деклараций...")
    
    declaration_files = []
    
    # Ищем все поддиректории с батчами
    batch_dirs = glob.glob(os.path.join(base_dir, "batch_*"))
    
    for batch_dir in sorted(batch_dirs, key=extract_date_from_filename):
        # Ищем все JSON файлы в каждой поддиректории (кроме файлов отчетов)
        json_files = [f for f in glob.glob(os.path.join(batch_dir, "*.json")) 
                     if not f.endswith("download_report.json")]
        
        # Добавляем все файлы в список
        declaration_files.extend(json_files)
    
    print(f"Найдено {len(declaration_files)} файлов деклараций")
    return declaration_files


def get_value_safely(data, *keys, default=""):
    """Безопасно извлекает значение из вложенного словаря по цепочке ключей"""
    current = data
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            return default
    
    # Преобразуем None в пустую строку
    if current is None:
        return default
    
    # Если получили список, возвращаем первый элемент или склеиваем все элементы
    if isinstance(current, list):
        if len(current) == 0:
            return default
        # Если список объектов, пытаемся извлечь полезную информацию из первого элемента
        elif isinstance(current[0], dict):
            if len(current) == 1:
                # Если в словаре есть понятное значение, возвращаем его
                for key in ['Name', 'Description', 'Text', 'Value', 'Id']:
                    if key in current[0]:
                        return str(current[0][key])
                # Иначе возвращаем весь словарь как строку
                return str(current[0])
            else:
                # При множестве элементов возвращаем список строковых представлений
                return ", ".join(str(item) for item in current)
        elif len(current) == 1:
            return str(current[0])
        else:
            return ", ".join(str(item) for item in current)
    
    return str(current)


def extract_declaration_data(file_path):
    """Извлекает необходимые данные из JSON файла декларации"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                print(f"Ошибка декодирования JSON в файле {file_path}")
                return None
        
        # Проверка валидности данных
        if not isinstance(data, dict):
            print(f"Неверный формат данных в файле {file_path}")
            return None
        
        # Находим основные данные документа
        doc_data = data.get("certdecltr_ConformityDocDetails", {})
        
        # Регистрационный номер
        doc_number = get_value_safely(doc_data, "DocId")
        
        # Ссылка на документ
        doc_id = data.get("certdecltr_id") or data.get("documents_id")
        doc_link = f"https://tsouz.belgiss.by/doc.php?id={doc_id}" if doc_id else ""
        
        # Статус действия сертификата
        doc_status = ""
        if "DocStatusDetails" in doc_data:
            status_code = get_value_safely(doc_data, "DocStatusDetails", "DocStatusCode")
            if status_code == "01":
                doc_status = "действует"
            elif status_code == "03":
                doc_status = "прекращен"
            elif status_code == "02":
                doc_status = "приостановлен"
        
        # Вид документа об оценке соответствия
        doc_type = ""
        conformity_kind_code = get_value_safely(doc_data, "ConformityDocKindCode")
        if conformity_kind_code == "10":
            doc_type = "Декларация о соответствии"
        elif conformity_kind_code == "1":
            doc_type = "Сертификат соответствия"
        
        # Номер технического регламента
        tech_regs = get_value_safely(doc_data, "TechnicalRegulationId", default=[])
        tech_reg = ", ".join(tech_regs) if isinstance(tech_regs, list) else tech_regs
        
        # Полное наименование органа по сертификации
        cert_org = get_value_safely(doc_data, "ConformityAuthorityV2Details", "BusinessEntityName")
        
        # Данные заявителя
        applicant = doc_data.get("ApplicantDetails", {})
        applicant_country = get_value_safely(applicant, "UnifiedCountryCode")
        applicant_name = get_value_safely(applicant, "BusinessEntityName")
        applicant_short_name = get_value_safely(applicant, "BusinessEntityBriefName")
        applicant_id = get_value_safely(applicant, "BusinessEntityId")
        
        # Адрес заявителя
        applicant_address = ""
        if "SubjectAddressDetails" in applicant and isinstance(applicant["SubjectAddressDetails"], list) and applicant["SubjectAddressDetails"]:
            addr = applicant["SubjectAddressDetails"][0]
            parts = [
                get_value_safely(addr, "RegionName"),
                get_value_safely(addr, "CityName"),
                get_value_safely(addr, "StreetName"),
                get_value_safely(addr, "BuildingNumberId")
            ]
            applicant_address = ", ".join(filter(None, parts))
        
        # Контактные данные заявителя
        applicant_contact = ""
        if "CommunicationDetails" in applicant and isinstance(applicant["CommunicationDetails"], list):
            contacts = []
            for comm in applicant["CommunicationDetails"]:
                if isinstance(comm, dict) and "CommunicationChannelId" in comm:
                    channel_id = comm["CommunicationChannelId"]
                    if isinstance(channel_id, list):
                        contacts.extend(channel_id)
                    else:
                        contacts.append(channel_id)
            applicant_contact = ", ".join(filter(None, contacts))
        
        # Данные изготовителя
        manufacturer = {}
        if "ManufacturerDetails" in doc_data and isinstance(doc_data["ManufacturerDetails"], list) and doc_data["ManufacturerDetails"]:
            manufacturer = doc_data["ManufacturerDetails"][0]
        
        manufacturer_country = get_value_safely(manufacturer, "UnifiedCountryCode")
        manufacturer_name = get_value_safely(manufacturer, "BusinessEntityBriefName")
        
        # Адрес изготовителя
        manufacturer_address = ""
        if "AddressV4Details" in manufacturer and isinstance(manufacturer["AddressV4Details"], list) and manufacturer["AddressV4Details"]:
            addr = manufacturer["AddressV4Details"][0]
            address_text = get_value_safely(addr, "AddressText")
            if address_text:
                manufacturer_address = address_text
            else:
                parts = [
                    get_value_safely(addr, "RegionName"),
                    get_value_safely(addr, "CityName"),
                    get_value_safely(addr, "StreetName"),
                    get_value_safely(addr, "BuildingNumberId")
                ]
                manufacturer_address = ", ".join(filter(None, parts))
        
        # Контактные данные изготовителя
        manufacturer_contact = ""
        if "CommunicationDetails" in manufacturer and isinstance(manufacturer["CommunicationDetails"], list):
            contacts = []
            for comm in manufacturer["CommunicationDetails"]:
                if isinstance(comm, dict) and "CommunicationChannelId" in comm:
                    channel_id = comm["CommunicationChannelId"]
                    if isinstance(channel_id, list):
                        contacts.extend(channel_id)
                    else:
                        contacts.append(channel_id)
            manufacturer_contact = ", ".join(filter(None, contacts))
        
        # Наименование объекта оценки соответствия
        product_name = get_value_safely(doc_data, "TechnicalRegulationObjectDetails", "ProductDetails", 0, "ProductName")
        
        # Код товара по ТН ВЭД ЕАЭС
        commodity_codes = get_value_safely(doc_data, "TechnicalRegulationObjectDetails", "ProductDetails", 0, "CommodityCode", default=[])
        commodity_code = ", ".join(commodity_codes) if isinstance(commodity_codes, list) else commodity_codes
        
        # Дополнительное наименование объекта
        product_text = get_value_safely(doc_data, "TechnicalRegulationObjectDetails", "ProductDetails", 0, "ProductText")
        
        # Даты документа
        doc_date = get_value_safely(doc_data, "DocStartDate")
        start_date = get_value_safely(doc_data, "DocStatusDetails", "StartDate") or doc_date
        end_date = get_value_safely(doc_data, "DocStatusDetails", "EndDate") or get_value_safely(doc_data, "DocValidityDate")
        
        return {
            "Регистрационный номер": doc_number,
            "Ссылка на документ": doc_link,
            "Статус действия сертификата (декларации)": doc_status,
            "Вид документа об оценке соответствия": doc_type,
            "Номер технического регламента": tech_reg,
            "Полное наименование органа по сертификации": cert_org,
            "Заявитель Страна": applicant_country,
            "Заявитель Краткое наименование": applicant_short_name,
            "Заявитель Идентификатор хозяйствующего субъекта": applicant_id,
            "Заявитель Адрес": applicant_address,
            "Заявитель Контактный реквизит": applicant_contact,
            "Изготовитель Страна": manufacturer_country,
            "Изготовитель Краткое наименование": manufacturer_name,
            "Изготовитель Адрес": manufacturer_address,
            "Изготовитель Контактный реквизит": manufacturer_contact,
            "Наименование объекта оценки соответствия": product_name,
            "Код товара по ТН ВЭД ЕАЭС": commodity_code,
            "Наименование объекта оценки соответствия (дополнительно)": product_text,
            "Дата документа": doc_date,
            "Дата начала действия": start_date,
            "Дата окончания действия": end_date
        }
        
    except Exception as e:
        print(f"Ошибка при обработке файла {file_path}: {str(e)}")
        return None


def find_processed_batches(output_dir="batch_results"):
    """Находит уже обработанные партии файлов"""
    if not os.path.exists(output_dir):
        return []
    
    # Ищем все файлы с шаблоном batch_NNNN.xlsx
    batch_files = glob.glob(os.path.join(output_dir, "batch_*.xlsx"))
    
    # Извлекаем номера партий из имен файлов
    batch_numbers = []
    for batch_file in batch_files:
        match = re.search(r'batch_(\d+)\.xlsx$', batch_file)
        if match:
            batch_numbers.append(int(match.group(1)))
    
    return sorted(batch_numbers)


def process_declaration_files(files, batch_size=1000, output_dir="batch_results", debug_mode=False, resume=False):
    """Обрабатывает файлы деклараций партиями и создает промежуточные файлы"""
    print("Обработка файлов деклараций...")
    
    # Устанавливаем глобальный флаг отладки
    global DEBUG_MODE
    DEBUG_MODE = debug_mode
    
    # Создаем директорию для результатов батчей, если не существует
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Создана директория для промежуточных результатов: {output_dir}")
    
    # Отладочный режим - обрабатываем только один файл и выводим его содержимое
    if debug_mode and files:
        print(f"ОТЛАДКА: Подробный анализ файла {files[0]}")
        try:
            with open(files[0], 'r', encoding='utf-8') as f:
                data = json.load(f)
                print(json.dumps(data, indent=2, ensure_ascii=False)[:1000] + "...")
                
                # Выводим ключи верхнего уровня
                print("\nКлючи верхнего уровня:", list(data.keys()))
                
                # Если есть certdecltr_ConformityDocDetails, выводим его ключи
                if "certdecltr_ConformityDocDetails" in data:
                    print("\nКлючи в certdecltr_ConformityDocDetails:", 
                          list(data["certdecltr_ConformityDocDetails"].keys()))
        except Exception as e:
            print(f"Ошибка при чтении файла: {str(e)}")
        
        # Обрабатываем только первый файл
        declaration_data = extract_declaration_data(files[0])
        if declaration_data:
            print("\nИзвлеченные данные:")
            for key, value in declaration_data.items():
                print(f"  {key}: {value}")
            
            # Создаем тестовый файл
            test_data = [declaration_data]
            test_file = os.path.join(output_dir, "test_batch.xlsx")
            create_report(test_data, test_file)
            return [test_file]
        else:
            print("Не удалось извлечь данные из тестового файла")
            return []
    
    # Разбиваем все файлы на батчи по batch_size
    total_files = len(files)
    batches = [files[i:i+batch_size] for i in range(0, total_files, batch_size)]
    
    print(f"Всего {total_files} файлов, разбито на {len(batches)} партий по {batch_size} файлов")
    
    # Если указан режим продолжения, находим уже обработанные партии
    processed_batch_numbers = []
    if resume:
        processed_batch_numbers = find_processed_batches(output_dir)
        if processed_batch_numbers:
            print(f"Найдено {len(processed_batch_numbers)} уже обработанных партий: {processed_batch_numbers}")
    
    # Список файлов с результатами батчей
    batch_result_files = []
    
    # Добавляем уже существующие файлы в список результатов
    for batch_num in processed_batch_numbers:
        batch_file = os.path.join(output_dir, f"batch_{batch_num:04d}.xlsx")
        if os.path.exists(batch_file):
            batch_result_files.append(batch_file)
            print(f"Добавлен ранее обработанный файл: {batch_file}")
    
    # Обрабатываем каждый батч
    for batch_num, batch_files in enumerate(batches):
        # Пропускаем уже обработанные партии
        batch_number = batch_num + 1
        if resume and batch_number in processed_batch_numbers:
            print(f"Пропуск уже обработанной партии {batch_number}/{len(batches)}")
            continue
        
        print(f"\nОбработка партии {batch_number}/{len(batches)} ({len(batch_files)} файлов)")
        
        # Список для хранения данных о декларациях в текущем батче
        batch_declarations_data = []
        
        # Счетчики для текущего батча
        processed_count = 0
        success_count = 0
        error_count = 0
        
        # Обрабатываем каждый файл в батче
        for i, file_path in enumerate(batch_files):
            # Извлекаем данные из файла
            declaration_data = extract_declaration_data(file_path)
            
            processed_count += 1
            
            # Если данные успешно извлечены, добавляем их в список
            if declaration_data:
                batch_declarations_data.append(declaration_data)
                success_count += 1
            else:
                error_count += 1
            
            # Отображаем прогресс каждые 100 файлов
            if (i + 1) % 100 == 0 or i + 1 == len(batch_files):
                print(f"  Прогресс: {i + 1}/{len(batch_files)} ({(i + 1) / len(batch_files) * 100:.1f}%)")
                print(f"    Успешно: {success_count}, Ошибок: {error_count}")
        
        # Если есть данные в текущем батче, сохраняем их во временный файл
        if batch_declarations_data:
            # Создаем имя файла для текущего батча
            batch_file = os.path.join(output_dir, f"batch_{batch_number:04d}.xlsx")
            
            print(f"  Сохранение результатов партии {batch_number} в файл {batch_file}...")
            
            # Создаем отчет для текущего батча
            create_report(batch_declarations_data, batch_file, "xlsx")
            
            # Добавляем файл в список результатов
            batch_result_files.append(batch_file)
            
            print(f"  Партия {batch_number} обработана: {success_count} записей")
        else:
            print(f"  ВНИМАНИЕ: Партия {batch_number} не содержит данных для сохранения")
    
    print(f"\nОбработка всех партий завершена. Создано {len(batch_result_files)} файлов с результатами.")
    return batch_result_files


def parse_date(date_str):
    """Пытается разобрать дату из строки в различных форматах"""
    if not date_str:
        return None
    
    # Если date_str не строка, преобразуем в строку или обрабатываем другие типы
    if not isinstance(date_str, str):
        if isinstance(date_str, (int, float)):
            try:
                # Возможно это число, которое Excel хранит как дату (числовое представление)
                # Преобразуем в datetime, который pandas может интерпретировать
                return pd.to_datetime(date_str, unit='D', origin='1899-12-30')
            except Exception:
                # Если преобразование не удалось, возвращаем строковое представление
                date_str = str(date_str)
        else:
            # Для других типов (включая datetime) просто возвращаем объект
            return date_str
    
    formats = [
        '%d.%m.%Y',  # 01.01.2020
        '%Y-%m-%d',  # 2020-01-01
        '%Y-%m-%dT%H:%M:%S',  # 2020-01-01T00:00:00
        '%Y/%m/%d',  # 2020/01/01
        '%d-%m-%Y',  # 01-01-2020
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    # Если не удалось разобрать дату стандартными форматами, пробуем извлечь ее из строки
    # Ищем шаблон ДД.ММ.ГГГГ
    match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', date_str)
    if match:
        try:
            day, month, year = map(int, match.groups())
            return datetime(year, month, day)
        except ValueError:
            pass
    
    return None


def merge_batch_results(batch_files, output_file="declarations_final.xlsx"):
    """Объединяет результаты обработки батчей в один итоговый файл"""
    if not batch_files:
        print("Нет файлов для объединения")
        return None
    
    print(f"\nОбъединение {len(batch_files)} файлов в итоговый отчет...")
    
    # Список для всех данных деклараций
    all_data = []
    
    # Загружаем данные из каждого батч-файла
    for i, batch_file in enumerate(batch_files):
        print(f"  Загрузка данных из файла {i+1}/{len(batch_files)}: {batch_file}")
        
        try:
            # Загружаем Excel или CSV в зависимости от расширения
            if batch_file.endswith('.xlsx'):
                # Читаем Excel без автоматического преобразования типов
                df = pd.read_excel(batch_file, parse_dates=False)
            elif batch_file.endswith('.csv'):
                df = pd.read_csv(batch_file, encoding='utf-8-sig')
            else:
                print(f"    Пропуск файла с неизвестным форматом: {batch_file}")
                continue
            
            # Убедимся, что колонка с датой обрабатывается как строка
            if 'Дата документа' in df.columns:
                df['Дата документа'] = df['Дата документа'].astype(str)
            
            print(f"    Загружено {len(df)} записей")
            
            # Преобразуем DataFrame в список словарей и добавляем в общий список
            batch_data = df.to_dict('records')
            all_data.extend(batch_data)
            
        except Exception as e:
            print(f"    Ошибка при загрузке файла {batch_file}: {str(e)}")
    
    print(f"Всего загружено {len(all_data)} записей из {len(batch_files)} файлов")
    
    # Создаем итоговый отчет
    if all_data:
        print(f"Создание итогового файла {output_file}...")
        final_file = create_report(all_data, output_file, "xlsx")
        print(f"Итоговый файл создан: {final_file}")
        return final_file
    else:
        print("Ошибка: не удалось загрузить данные из промежуточных файлов")
        return None


def create_report(declarations_data, output_file="declarations_report", format_type="xlsx"):
    """Создает отчет с данными деклараций в указанном формате"""
    if not declarations_data:
        print("Ошибка: нет данных для создания отчета")
        return None
    
    print(f"Создание отчета в формате {format_type}...")
    
    # Получаем имя файла без расширения
    output_base = os.path.splitext(output_file)[0]
    
    # Добавляем правильное расширение
    if format_type == "xlsx":
        full_output_file = f"{output_base}.xlsx"
    else:  # csv по умолчанию
        full_output_file = f"{output_base}.csv"
    
    print(f"Подготовка {len(declarations_data)} записей...")
    
    # Создаем DataFrame из данных деклараций
    df = pd.DataFrame(declarations_data)
    
    # Предварительная обработка для отладки
    print(f"Размер данных: {df.shape}")
    print(f"Столбцы: {df.columns.tolist()}")
    
    # Подсчет непустых значений в каждом столбце
    non_empty_counts = {col: df[col].astype(str).str.strip().str.len().gt(0).sum() 
                       for col in df.columns}
    print("Количество непустых значений в столбцах:")
    for col, count in non_empty_counts.items():
        print(f"  {col}: {count} из {len(df)} ({count/len(df)*100:.1f}%)")
    
    # Преобразуем даты для сортировки
    print("Подготовка данных для сортировки...")
    
    # Создаем новый столбец для хранения дат в формате datetime
    df['Дата_сортировки'] = None
    
    # Счетчики для отслеживания преобразования дат
    date_conversion_success = 0
    date_conversion_failures = 0
    
    # Извлекаем даты из строк и преобразуем их в объекты datetime
    for i, row in df.iterrows():
        date_str = row.get('Дата документа', '')
        parsed_date = parse_date(date_str)
        
        if parsed_date:
            df.at[i, 'Дата_сортировки'] = parsed_date
            date_conversion_success += 1
        else:
            date_conversion_failures += 1
            # Для записей без даты используем минимальную дату
            df.at[i, 'Дата_сортировки'] = datetime.min
    
    print(f"Преобразование дат: успешно {date_conversion_success}, не удалось {date_conversion_failures}")
    
    if date_conversion_success > 0:
        print("Сортировка данных по дате...")
        # Сортируем данные по дате
        df = df.sort_values(by='Дата_сортировки')
        print("Сортировка завершена")
    else:
        print("ВНИМАНИЕ: не удалось извлечь ни одной даты для сортировки")
    
    # Удаляем служебные столбцы
    if 'Дата_сортировки' in df.columns:
        df = df.drop(columns=['Дата_сортировки'])
    
    # Сохраняем отчет в указанном формате
    print(f"Сохранение отчета в файл: {full_output_file}")
    
    if format_type == "xlsx":
        try:
            # Пробуем сохранить в Excel
            df.to_excel(full_output_file, index=False)
            
            # Форматируем Excel-файл
            format_excel_file(full_output_file)
        except Exception as e:
            print(f"Ошибка при сохранении в Excel: {str(e)}")
            
            # В случае ошибки сохраняем в CSV как запасной вариант
            csv_backup = f"{output_base}_backup.csv"
            print(f"Пробуем сохранить в CSV: {csv_backup}")
            df.to_csv(csv_backup, index=False, encoding='utf-8-sig')
            
            return csv_backup
    else:
        # Сохраняем в CSV
        df.to_csv(full_output_file, index=False, encoding='utf-8-sig')
    
    print(f"Отчет успешно создан: {full_output_file}")
    return full_output_file


def format_excel_file(file_path):
    """Форматирует Excel-файл: устанавливает ширину колонок и фиксирует первые строки"""
    print(f"Настройка форматирования Excel-файла: {file_path}")
    
    try:
        # Загружаем файл
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Устанавливаем выравнивание для всех ячеек
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Устанавливаем автоширину для всех колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Проверяем длину текста в каждой ячейке колонки
            for cell in column[:min(500, len(column))]:  # Проверяем до 500 строк для определения ширины (для скорости)
                if cell.value:
                    try:
                        cell_text = str(cell.value)
                        
                        # Для заголовков увеличиваем ширину
                        if cell.row == 1:
                            text_length = len(cell_text) * 1.3
                            # Минимальная ширина для заголовков
                            text_length = max(text_length, 15)
                        else:
                            # Для обычных ячеек - зависит от длины содержимого
                            # Учитываем переносы строк
                            if "\n" in cell_text:
                                lines = cell_text.split("\n")
                                text_length = max(len(line) for line in lines) * 1.1
                            else:
                                text_length = len(cell_text) * 1.1
                        
                        max_length = max(max_length, text_length)
                    except Exception:
                        # В случае ошибки пропускаем ячейку
                        continue
            
            # Устанавливаем ширину колонки (с ограничением минимальной и максимальной ширины)
            # Минимальная ширина - 10, максимальная - 80
            adjusted_width = min(max(max_length + 2, 10), 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Создаем гиперссылки в колонке "Ссылка на документ"
        link_column_index = None
        for col_idx, cell in enumerate(ws[1], 1):  # Индексы колонок в openpyxl начинаются с 1
            if cell.value == "Ссылка на документ":
                link_column_index = col_idx
                break
        
        if link_column_index:
            link_column_letter = get_column_letter(link_column_index)
            # Проходим по всем строкам, начиная со второй (пропускаем заголовок)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws[f"{link_column_letter}{row_idx}"]
                if cell.value and isinstance(cell.value, str) and (cell.value.startswith('http://') or cell.value.startswith('https://')):
                    # Создаем гиперссылку
                    from openpyxl.styles import Font
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")  # Синий цвет и подчеркивание для ссылок
        
        # Фиксируем первую строку
        ws.freeze_panes = 'A2'
        
        # Сохраняем отформатированный файл
        wb.save(file_path)
        print("Форматирование Excel-файла завершено успешно")
    except Exception as e:
        print(f"Ошибка при форматировании Excel-файла: {str(e)}")
        # Продолжаем выполнение скрипта даже при ошибке форматирования


def main():
    # Парсер аргументов командной строки
    parser = argparse.ArgumentParser(description='Генерация отчета на основе JSON-файлов деклараций')
    parser.add_argument('--source-dir', type=str, default='declarations_details', 
                        help='Директория с данными деклараций (по умолчанию: declarations_details)')
    parser.add_argument('--output', type=str, default='declarations_report.xlsx', 
                        help='Имя итогового Excel-файла (по умолчанию: declarations_report.xlsx)')
    parser.add_argument('--batch-dir', type=str, default='batch_results', 
                        help='Директория для промежуточных файлов (по умолчанию: batch_results)')
    parser.add_argument('--batch-size', type=int, default=1000, 
                        help='Размер партии для обработки (по умолчанию: 1000)')
    parser.add_argument('--debug', action='store_true', help='Режим отладки (обработка только одного файла)')
    parser.add_argument('--test-file', type=str, help='Путь к конкретному файлу для тестирования')
    parser.add_argument('--merge-only', action='store_true', 
                        help='Только объединить существующие промежуточные файлы без обработки исходных данных')
    parser.add_argument('--resume', action='store_true',
                       help='Продолжить обработку с места последней остановки, пропуская уже обработанные партии')

    args = parser.parse_args()

    try:
        # Тестирование одного конкретного файла
        if args.test_file:
            print(f"Тестирование файла: {args.test_file}")
            if os.path.exists(args.test_file):
                process_declaration_files([args.test_file], output_dir=args.batch_dir, debug_mode=True)
                return
            else:
                print(f"Ошибка: файл {args.test_file} не найден")
                return
        
        # Если указан режим объединения промежуточных файлов, то только объединяем их
        if args.merge_only:
            if os.path.exists(args.batch_dir):
                batch_files = glob.glob(os.path.join(args.batch_dir, "batch_*.xlsx"))
                if batch_files:
                    merge_batch_results(batch_files, args.output)
                else:
                    print(f"В директории {args.batch_dir} не найдены промежуточные файлы")
            else:
                print(f"Директория {args.batch_dir} не существует")
            return
        
        # Находим все файлы деклараций
        declaration_files = find_declaration_files(args.source_dir)
        
        if not declaration_files:
            print(f"Ошибка: не найдены файлы деклараций в директории {args.source_dir}")
            return
        
        # Обрабатываем файлы и создаем промежуточные файлы
        batch_files = process_declaration_files(
            declaration_files, 
            batch_size=args.batch_size, 
            output_dir=args.batch_dir,
            debug_mode=args.debug,
            resume=args.resume
        )
        
        if not batch_files:
            print("Ошибка: не удалось создать файлы с результатами обработки партий")
            return
        
        # Объединяем результаты в итоговый файл
        final_file = merge_batch_results(batch_files, args.output)
        
        if final_file:
            print(f"Обработка завершена. Итоговый отчет сохранен в файле: {final_file}")
        else:
            print("Ошибка: не удалось создать итоговый отчет")
    
    except Exception as e:
        print(f"Произошла критическая ошибка: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    start_time = datetime.now()
    print(f"Начало выполнения: {start_time}")
    
    try:
        main()
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")
    
    end_time = datetime.now()
    duration = end_time - start_time
    print(f"Время выполнения: {duration}") 